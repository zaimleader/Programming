from openpyxl.styles import *
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

path = "Python/Pandas/resources/excel/"

wb = Workbook()
ws= wb.active

for i in range(1, 20):
    ws.append(range(30))

ws.merge_cells("A1:B5")
ws.unmerge_cells("A1:B5")
ws.merge_cells(start_row=2, start_column=2, end_row=5, end_column=5)

cell = ws["B2"]

cell.font = Font(color="FF58FF", size=40, italic=True, bold=True)
cell.value = "Merged Cell"
cell.alignment = Alignment(horizontal="right", vertical="bottom")
cell.fill = GradientFill(stop=("1058FF", "F37EA1"))

# wb.save(path + "save/stylingSheet.xlsx")

heighlight = NamedStyle(name="heighlight")

heighlight.font = Font(bold=True)
bd = Side(style="thick", color="FF0000")
heighlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
heighlight.fill = PatternFill("solid", fgColor="0000FF")

count = 0
for col in ws.iter_cols(min_col=8, min_row=1, max_col=30, max_row=30):
    col[count].style = heighlight
    count = count + 1

wb.save(path + "save/heighlight.xlsx")