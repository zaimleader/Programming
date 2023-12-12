from openpyxl.workbook import Workbook
from openpyxl import load_workbook

path = "Python/Pandas/resources/excel/"

wb = load_workbook(path + "regions.xlsx")
ws = wb.active

cell_rang = ws["A1":"C1"]
print("------- cell range -------")
print(cell_rang)
print("\n")

cols_range = ws["A":"D"]
print("------- cols range -------")
print(cols_range)
print("\n")

rows_range = ws[1:3]
print("------- rows range -------")
print(rows_range)
print("\n")

print("------- display data with range -------")
for row in ws.iter_rows(min_row=2, max_col=4, max_row=5, values_only=True):
    for cell in row:
        print(cell)
    print("\n")