from openpyxl.workbook import Workbook
from openpyxl import load_workbook

path = "Python/Pandas/resources/excel/"

wb = Workbook()
ws = wb.active
ws.title = "Books"

ws1 = wb.create_sheet("Note Book")
ws2 = wb.create_sheet("Journal")

print(wb.sheetnames)

wb.save(path + "save/books.xlsx")

wb2 = load_workbook(path + "regions.xlsx")
work_sheet_active = wb2.active

new_work_sheet = wb2.create_sheet("cities", 0)

work_sheet_active["A1"] = "something" #if you xant to change title A1 by another title

cell = work_sheet_active["A1"]

print(cell.value)
